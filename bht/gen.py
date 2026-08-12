import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
d=json.load(open('extracted.json')); comp=json.load(open('computed.json'))
site=d['site']; staff=d['staff']; roster=d['roster']
days=list(range(1,31))
# roster lookup: (staff,day)->pos
rmap={}
for e in roster: rmap[(e['staff'],e['day'])]=e['pos']
FONT='Arial'; thin=Side(style='thin',color='BBBBBB')
border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='1F3864'); hdrfont=Font(name=FONT,bold=True,color='FFFFFF',size=10)
subfill=PatternFill('solid',fgColor='D9E1F2'); vacfill=PatternFill('solid',fgColor='FCE4EC')
def style_hdr(c): c.fill=hdrfill; c.font=hdrfont; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=border
def base(p):
    f=p.split('/')[0].strip()
    for k in ['責','日B','日C','夜A','夜B']:
        if f.startswith(k): return k
    if f.startswith('臨'): return '臨時'
    if '研' in f: return '研'
    return f
POS=['責','日B','日C','夜A','夜B','臨時']

# ---------- 配置表.xlsx (staff × day, actual names) ----------
def build_haichi(path, title, planned=False):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='配置表'
    ws['A1']=title; ws['A1'].font=Font(name=FONT,bold=True,size=14)
    ws['A2']=f"現場名：{site['name']}　業務種別：{site['gyoumu']}　受注形態：{site['jyuchu']}　所属営業所：{site['office']}"
    ws['A2'].font=Font(name=FONT,size=10)
    ws['A3']=f"対象期間：{site['year']}年{site['month']}月1日 ～ {site['month']}月{site['days']}日"
    ws['A3'].font=Font(name=FONT,size=10)
    r0=5
    # header row: No | 氏名 | day1..30 | 勤務数
    ws.cell(r0,1,'No'); ws.cell(r0,2,'警備員氏名')
    for i,dy in enumerate(days): ws.cell(r0,3+i,dy)
    ws.cell(r0,3+len(days),'勤務数')
    for c in range(1,4+len(days)): style_hdr(ws.cell(r0,c))
    # position headcount summary block (COUNTIF formulas) above? put below staff.
    # staff rows
    rr=r0+1
    for idx,s in enumerate(staff,1):
        ws.cell(rr,1,idx).font=Font(name=FONT,size=9)
        cell=ws.cell(rr,2,s); cell.font=Font(name=FONT,size=9); cell.border=border
        ws.cell(rr,1).border=border; ws.cell(rr,1).alignment=Alignment(horizontal='center')
        for i,dy in enumerate(days):
            pos=rmap.get((s,dy))
            cc=ws.cell(rr,3+i, pos if pos else None)
            cc.font=Font(name=FONT,size=9); cc.border=border; cc.alignment=Alignment(horizontal='center')
        # 勤務数 = COUNTA of day cells
        first=get_column_letter(3); last=get_column_letter(2+len(days))
        wc=ws.cell(rr,3+len(days)); wc.value=f"=COUNTA({first}{rr}:{last}{rr})"; wc.font=Font(name=FONT,size=9,bold=True); wc.border=border; wc.alignment=Alignment(horizontal='center')
        rr+=1
    # summary: per-position per-day headcount (COUNTIF over staff block, matching base code via wildcard)
    sumr=rr+1
    ws.cell(sumr,2,'■ 区分別 配置人数（自動集計）').font=Font(name=FONT,bold=True,size=10)
    sumr+=1
    ws.cell(sumr,2,'勤務区分'); 
    for i,dy in enumerate(days): ws.cell(sumr,3+i,dy)
    ws.cell(sumr,3+len(days),'延べ')
    for c in range(2,4+len(days)): style_hdr(ws.cell(sumr,c))
    staff_first=r0+1; staff_last=r0+len(staff)
    for j,p in enumerate(POS):
        row=sumr+1+j
        lc=ws.cell(row,2,p); lc.font=Font(name=FONT,size=9,bold=True); lc.fill=subfill; lc.border=border; lc.alignment=Alignment(horizontal='center')
        for i,dy in enumerate(days):
            col=get_column_letter(3+i)
            # count cells in this day column (staff rows) whose base code == p → use COUNTIF with wildcard prefix
            pat = "*"+p+"*"
            f=f'=COUNTIF({col}{staff_first}:{col}{staff_last},"{pat}")'
            cc=ws.cell(row,3+i); cc.value=f; cc.font=Font(name=FONT,size=9); cc.border=border; cc.alignment=Alignment(horizontal='center')
        tot=ws.cell(row,3+len(days)); tot.value=f'=SUM({get_column_letter(3)}{row}:{get_column_letter(2+len(days))}{row})'; tot.font=Font(name=FONT,size=9,bold=True); tot.border=border; tot.alignment=Alignment(horizontal='center')
    # 計 row (sum of positions per day)
    calc=sumr+1+len(POS)
    lc=ws.cell(calc,2,'計'); lc.font=Font(name=FONT,size=9,bold=True); lc.fill=PatternFill('solid',fgColor='FFF2CC'); lc.border=border; lc.alignment=Alignment(horizontal='center')
    for i,dy in enumerate(days):
        col=get_column_letter(3+i)
        cc=ws.cell(calc,3+i); cc.value=f'=SUM({col}{sumr+1}:{col}{calc-1})'; cc.font=Font(name=FONT,size=9,bold=True); cc.border=border; cc.alignment=Alignment(horizontal='center')
    ws.cell(calc,3+len(days)).value=f'=SUM({get_column_letter(3)}{calc}:{get_column_letter(2+len(days))}{calc})'
    ws.cell(calc,3+len(days)).font=Font(name=FONT,size=9,bold=True); ws.cell(calc,3+len(days)).border=border
    # legend
    lg=calc+2; ws.cell(lg,2,'（凡例）勤務区分の時間帯').font=Font(name=FONT,bold=True,size=10)
    for k,l in enumerate(d['position_legend']):
        ws.cell(lg+1+k,2,l['code']).font=Font(name=FONT,size=9)
        ws.cell(lg+1+k,3,l['time']).font=Font(name=FONT,size=9)
    # widths
    ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=12
    for i in range(len(days)): ws.column_dimensions[get_column_letter(3+i)].width=6.5
    ws.column_dimensions[get_column_letter(3+len(days))].width=7
    ws.freeze_panes='C6'
    wb.save(path); print('saved',path)

build_haichi('out/配置表_2026年06月_ブルガリホテル東京.xlsx','配置表')
build_haichi('out/配置予定表_2026年06月_ブルガリホテル東京.xlsx','配置予定表', planned=True)
