import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
d=json.load(open('extracted.json'))
site=d['site']; dc=d['daily_counts']; tot=d['totals']
FONT='Arial'; thin=Side(style='thin',color='BBBBBB'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='1F3864'); hdrfont=Font(name=FONT,bold=True,color='FFFFFF',size=10)
subfill=PatternFill('solid',fgColor='D9E1F2')
def sh(c): c.fill=hdrfill; c.font=hdrfont; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=border
# categories to report (exclude 天気日 noise; keep meaningful counters)
cats=['入館者','外部スタッフ','巡回時未施錠','警察対応','自火報発報','ジュエリーケース発報','救急対応','不審物対応','不審者対応','エレベーター呼出','緊急呼出','未返却','誤進入','セキュリティカード登録・変更']
wb=openpyxl.Workbook()
# --- data sheet: per-day matrix ---
ds=wb.active; ds.title='日別集計'
ds.cell(1,1,'日'); 
for j,cat in enumerate(cats): ds.cell(1,2+j,cat)
ds.cell(1,2+len(cats),'稼働率')
for c in range(1,3+len(cats)): sh(ds.cell(1,c))
for i,row in enumerate(dc):
    ds.cell(2+i,1,row['day']).border=border
    for j,cat in enumerate(cats):
        cc=ds.cell(2+i,2+j,row.get(cat,0)); cc.border=border; cc.font=Font(name=FONT,size=9); cc.alignment=Alignment(horizontal='center')
    rc=ds.cell(2+i,2+len(cats), round(row['稼働率'],4) if row.get('稼働率') is not None else 0); rc.number_format='0.0%'; rc.border=border; rc.font=Font(name=FONT,size=9)
# totals row (SUM formulas)
tr=2+len(dc)
ds.cell(tr,1,'合計').font=Font(name=FONT,bold=True); ds.cell(tr,1).fill=subfill; ds.cell(tr,1).border=border
for j,cat in enumerate(cats):
    col=get_column_letter(2+j); f=f'=SUM({col}2:{col}{tr-1})'
    cc=ds.cell(tr,2+j); cc.value=f; cc.font=Font(name=FONT,bold=True,size=9); cc.fill=subfill; cc.border=border; cc.alignment=Alignment(horizontal='center')
# avg utilization
col=get_column_letter(2+len(cats)); ac=ds.cell(tr,2+len(cats)); ac.value=f'=AVERAGE({col}2:{col}{tr-1})'; ac.number_format='0.0%'; ac.font=Font(name=FONT,bold=True,size=9); ac.fill=subfill; ac.border=border
ds.column_dimensions['A'].width=5
for j in range(len(cats)): ds.column_dimensions[get_column_letter(2+j)].width=11
ds.freeze_panes='B2'

# --- report sheet ---
rs=wb.create_sheet('月次報告書',0)
rs['A1']='警 備 月 報'; rs['A1'].font=Font(name=FONT,bold=True,size=16)
rs['A3']=f"【{site['name']}】"; rs['A3'].font=Font(name=FONT,bold=True,size=12)
rs['A4']=f"対象期間：{site['year']}年{site['month']}月1日 ～ {site['month']}月{site['days']}日　／　所属営業所：{site['office']}"
rs['A4'].font=Font(name=FONT,size=10)
rs['A6']='■ 業務対応件数等 報告事項'; rs['A6'].font=Font(name=FONT,bold=True,size=11)
r=7
rs.cell(r,1,'項目'); rs.cell(r,2,'月間件数'); rs.cell(r,3,'単位')
for c in range(1,4): sh(rs.cell(r,c))
tr_data=2+len(dc)  # totals row index on 日別集計
for k,cat in enumerate(cats):
    row=r+1+k
    lc=rs.cell(row,1,cat); lc.font=Font(name=FONT,size=10); lc.border=border
    col=get_column_letter(2+k)
    vc=rs.cell(row,2); vc.value=f"='日別集計'!{col}{tr_data}"; vc.font=Font(name=FONT,size=10,bold=True); vc.border=border; vc.alignment=Alignment(horizontal='center')
    unit='名' if cat in ('入館者','外部スタッフ') else '件'
    uc=rs.cell(row,3,unit); uc.font=Font(name=FONT,size=10); uc.border=border; uc.alignment=Alignment(horizontal='center')
after=r+1+len(cats)
rs.cell(after+1,1,'平均稼働率'); rs.cell(after+1,1).font=Font(name=FONT,bold=True)
avgcol=get_column_letter(2+len(cats))
ac=rs.cell(after+1,2); ac.value=f"='日別集計'!{avgcol}{tr_data}"; ac.number_format='0.0%'; ac.font=Font(name=FONT,bold=True); ac.border=border
# incident narrative section
ns=after+3
rs.cell(ns,1,'■ 自火報・緊急対応 警備対応事案（明細）').font=Font(name=FONT,bold=True,size=11)
narr=[
 ('自火報発報　1件',''),
 ('6/3','光電アナログ注意発報（41階7区36番）。異常なし。'),
 ('緊急呼出　6件',''),
 ('6/4','緊急呼出発報(40F レストランWC2)。異常なし。'),
 ('6/5','緊急呼出発報（客室4110号室）ゲストによる誤操作。'),
 ('6/10','緊急呼出発報（客室4101号室）HSKP対応。'),
 ('6/12','緊急呼出発報(40階女性用ハマム・サウナ)。異常なし。'),
 ('6/22','緊急呼出発報（4101号室）FDへ連絡。'),
 ('6/24','緊急呼出発報（客室4110号室）ゲスト誤操作。'),
 ('不審者対応　4件',''),
 ('6/19','45F BARに不審ゲスト。商業側より退館を確認。'),
 ('6/20','不審ゲスト対応。Duty/LP連携で対応。'),
 ('6/22','ITV監視依頼。翌日支払対応。'),
]
for k,(a,b) in enumerate(narr):
    rr=ns+1+k
    ca=rs.cell(rr,1,a); ca.font=Font(name=FONT,bold=(b==''),size=10)
    cb=rs.cell(rr,2,b); cb.font=Font(name=FONT,size=9)
rs.cell(ns+2+len(narr),1,'※ 明細テキストは日報の特記事項から自動集約（デモ）。実データ連携時は日報レコードから自動生成。').font=Font(name=FONT,italic=True,size=8,color='888888')
rs.column_dimensions['A'].width=26; rs.column_dimensions['B'].width=60; rs.column_dimensions['C'].width=6
wb.save('out/月次報告書_2026年06月_ブルガリホテル東京.xlsx'); print('saved 月次報告書')
